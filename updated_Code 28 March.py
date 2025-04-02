import pandas as pd
import datetime
import numpy as np
import mysql.connector as msql
from mysql.connector import Error
from google.cloud import bigquery

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import os 
os.chdir(r"D:\Oushnik Sarkar\Python")
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'D:/Oushnik Sarkar/data-warehousing-prod.json'

#Imports google cloud client library and initiates BQ service
from google.cloud import bigquery
#from google.cloud import datastore
bigquery_client = bigquery.Client()

QUERY = """
SELECT
    Centre, FinYear,Season, SaleNo, AuctionDate, LotNo, 
Garden, GardenMDM,Grade, GradeMDM, InvoiceNo, 
Buyer, BuyerMDM, BuyerGroup, BrokerCode,
Seller, SellerGroup, Category, SubCategory, TeaType, 
SubTeaType,LotStatus, Area, EstBlf,GPDATE,ReprintNo,
SUM(IF(LotStatus = 'Sold',TotalWeight,InvoiceWeight)) AS Offer_Qty,
SUM(TotalWeight) AS Sold_Qty,
SUM(Value) AS Total_Value

FROM `data-warehousing-prod.EasyReports.SaleTransactionView`

WHERE Season IN (2024) And SellerGroup = "LUXMI"

GROUP BY Centre, FinYear,Season, SaleNo, AuctionDate, LotNo, Garden, GardenMDM,Grade, GradeMDM, 
InvoiceNo, Buyer, BuyerMDM, BuyerGroup, BrokerCode,Seller, SellerGroup, Category, SubCategory, 
TeaType, SubTeaType, LotStatus, Area, EstBlf,GPDATE,ReprintNo """

Query_Results = bigquery_client.query(QUERY)
df = Query_Results.to_dataframe()

df['Avg_Price'] = df['Total_Value'] / df['Sold_Qty']
df['SaleAlies'] = np.where((df['SaleNo'] >= 1) & (df['SaleNo'] <= 13), df['SaleNo'] + 52, df['SaleNo'])

df.info()

#-----------------------------------------------UPTO SALE-------------------------------------------------------

df1=df[(df['Category'].isin(["CTC"])) & (df['EstBlf']=="EST") & (df['SaleAlies'].between(14,60))]

# Step 1: Aggregate Data Before Pivoting
summary_df = (df1.groupby(["SubTeaType", "GradeMDM", "GardenMDM"]).agg({
    "Offer_Qty":"sum","Sold_Qty":"sum","Avg_Price":"mean"}).reset_index())

# Step 2: Create Pivot Table with Multi-Index
pivot_df = summary_df.pivot_table(
    index=["SubTeaType", "GradeMDM"],  # Multi-index (SubTeaType -> GradeMDM)
    columns="GardenMDM",  # Columns
    values=["Offer_Qty","Sold_Qty", "Avg_Price"],  # Metrics to show
    fill_value=0,  # Replace NaNs with 0
    aggfunc={"Offer_Qty":"sum","Sold_Qty": "sum", "Avg_Price": "mean"}).round(0)

#SWAPING
pivot_df=pivot_df.swaplevel(axis=1).sort_index(axis=1)

#################Creating Percentage#################
sold_qty = pivot_df.xs('Sold_Qty', axis=1, level=1)
total_sold_qty = sold_qty.sum()

# Calculate the percentage of parent total
pct_of_parent = sold_qty.divide(total_sold_qty, axis=1) * 100

# Add the percentage of parent total to the original DataFrame
for garden in pct_of_parent.columns:
    pivot_df[(garden, 'Grade%')] = pct_of_parent[garden]

# Sort columns for better readability
pivot_df = pivot_df.sort_index(axis=1)

#checking the performance of percentage
pivot_df.loc[:, pivot_df.columns.get_level_values(1) == 'Grade%'].sum()

#################Creating Out%#################

offer_qty = pivot_df.xs('Offer_Qty', axis=1, level=1)
out_percentage = (1 - (sold_qty / offer_qty)) * 100

# Handle division by zero (if Offer_Qty is 0)
out_percentage = out_percentage.fillna(0)

# Add 'Out%' back to pivot_df
for garden in out_percentage.columns:
    pivot_df[(garden, 'Out%')] = out_percentage[garden]

# Sort columns for better readability
pivot_df = pivot_df.sort_index(axis=1)

# Checking the performance of 'Out%'
pivot_df.loc[:, pivot_df.columns.get_level_values(1) == 'Out%'].mean()

################################################
sum_cols = ['Sold_Qty', 'Grade%']
avg_cols = ['Avg_Price']
weight_col = 'Sold_Qty'

def add_subtotals(df):
    subtotals = []
    
    for category in ['PRIMARY', 'SECONDARY']:
        category_rows = df.loc[category]
        subtotal = pd.DataFrame(index=[(category, 'Subtotal')], columns=df.columns)
        
        for garden in df.columns.levels[0]:  
            # Sum Sold_Qty & Grade%
            subtotal[(garden, 'Sold_Qty')] = category_rows[(garden, 'Sold_Qty')].sum()
            subtotal[(garden, 'Grade%')] = category_rows[(garden, 'Grade%')].sum()
            
            # Weighted Average for Avg_Price
            total_weight = category_rows[(garden, weight_col)].sum()
            if total_weight > 0:
                subtotal[(garden, 'Avg_Price')] = (
                    (category_rows[(garden, 'Avg_Price')] * category_rows[(garden, weight_col)]).sum() / total_weight
                )
            else:
                subtotal[(garden, 'Avg_Price')] = 0

            # Sum Offer_Qty (Needed for Out% Calculation)
            subtotal[(garden, 'Offer_Qty')] = category_rows[(garden, 'Offer_Qty')].sum()

            # Correct Out% Calculation
            if subtotal[(garden, 'Offer_Qty')].iloc[0] > 0:
                subtotal[(garden, 'Out%')]=(1 -(subtotal[(garden,'Sold_Qty')] / subtotal[(garden,'Offer_Qty')]))*100
            else:
                subtotal[(garden, 'Out%')] = 0  # Avoid division by zero
        
        subtotals.append(subtotal)

    df = pd.concat([df] + subtotals).sort_index()

    # Compute Grand Total
    grand_total = pd.DataFrame(index=[('Grand Total', '')], columns=df.columns)

    for garden in df.columns.levels[0]:  
        grand_total[(garden, 'Sold_Qty')] = df.xs('Subtotal', level=1).loc[:, (garden, 'Sold_Qty')].sum()
        grand_total[(garden, 'Grade%')] = df.xs('Subtotal', level=1).loc[:, (garden, 'Grade%')].sum()
        grand_total[(garden, 'Offer_Qty')] = df.xs('Subtotal', level=1).loc[:, (garden, 'Offer_Qty')].sum()

        total_weight = df.xs('Subtotal', level=1).loc[:, (garden, weight_col)].sum()
        if total_weight > 0:
            grand_total[(garden, 'Avg_Price')] = (
                (df.xs('Subtotal', level=1).loc[:, (garden, 'Avg_Price')] * df.xs('Subtotal', level=1).loc[:, (garden, weight_col)]).sum()
                / total_weight
            )
        else:
            grand_total[(garden, 'Avg_Price')] = 0
        
        # Correct Out% Calculation for Grand Total
        if grand_total[(garden, 'Offer_Qty')].iloc[0] > 0:
            grand_total[(garden, 'Out%')] = (1 - (grand_total[(garden, 'Sold_Qty')] / grand_total[(garden, 'Offer_Qty')])) * 100
        else:
            grand_total[(garden, 'Out%')] = 0

    df = pd.concat([df, grand_total])
    
    return df

# Apply function
pivot_df = add_subtotals(pivot_df)

# Custom sorting function

desired_order = ['Sold_Qty', 'Grade%', 'Avg_Price', 'Out%']

# Rearrange columns under each Garden name
new_columns = []
for garden in pivot_df.columns.levels[0]:  # Iterate over the garden names
    for metric in desired_order:  # Maintain the desired metric order
        new_columns.append((garden, metric))

# Update the DataFrame with the new column order
pivot_df = pivot_df[new_columns]

#########################Grade Sequence####################
desired_order2 = [
    "BOPL", "BPS", "BOP", "BOPSM", "BP", "OF", "PF", "PD", "D", "CD",
    "BOPL1", "BPS1", "BOP1", "BOPSM1", "BP1", "OF1","PF1", "PD1", "D1", "CD1"]

def reorder_group(group):
    # Split "Subtotal" rows from others
    subtotal = group[group.index.get_level_values(1) == 'Subtotal']
    others = group[group.index.get_level_values(1) != 'Subtotal']
       
    others = others.reindex(
        sorted(others.index,
            key=lambda x: desired_order2.index(x[1]) if x[1] in desired_order2 else len(desired_order2)
        ))
    
    # Append "Subtotal" back at the end
    return pd.concat([others, subtotal])

# Separate "Grand Total" row
grand_total = pivot_df[pivot_df.index.get_level_values(0) == 'Grand Total']
pivot_df_without_grand_total = pivot_df[pivot_df.index.get_level_values(0) != 'Grand Total']

pivot_df_reordered = pivot_df_without_grand_total.groupby(level=0, group_keys=False).apply(reorder_group)

# Append "Grand Total" back at the bottom
pivot_df_final = pd.concat([pivot_df_reordered, grand_total])

#-----------------------------------------------FOR SALE-------------------------------------------------------

df_for = df[(df['Category'].isin(["CTC"])) & (df['EstBlf']=="EST") & (df['SaleAlies']==60)]

# Step 1: Aggregate Data Before Pivoting
summary_df2 = (df_for.groupby(["SubTeaType", "GradeMDM", "GardenMDM"]).agg({
    "Offer_Qty":"sum","Sold_Qty":"sum","Avg_Price":"mean"}).reset_index())

# Step 2: Create Pivot Table with Multi-Index
pivot_df2 = summary_df2.pivot_table(
    index=["SubTeaType", "GradeMDM"],  # Multi-index (SubTeaType -> GradeMDM)
    columns="GardenMDM",  # Columns
    values=["Offer_Qty","Sold_Qty", "Avg_Price"],  # Metrics to show
    fill_value=0,  # Replace NaNs with 0
    aggfunc={"Offer_Qty":"sum","Sold_Qty": "sum", "Avg_Price": "mean"}).round(0)

#SWAPING
pivot_df2 = pivot_df2.swaplevel(axis=1).sort_index(axis=1)

#################Creating Percentage#################
sold_qty2 = pivot_df2.xs('Sold_Qty', axis=1, level=1)
total_sold_qty2 = sold_qty2.sum()

# Calculate the percentage of parent total
pct_of_parent2 = sold_qty2.divide(total_sold_qty2, axis=1) * 100

# Add the percentage of parent total to the original DataFrame
for garden in pct_of_parent2.columns:
    pivot_df2[(garden, 'Grade%')] = pct_of_parent2[garden]

# Sort columns for better readability
pivot_df2 = pivot_df2.sort_index(axis=1)

#checking the performance of percentage
pivot_df2.loc[:, pivot_df2.columns.get_level_values(1) == 'Grade%'].sum()

#################Creating Out%#################

offer_qty2 = pivot_df2.xs('Offer_Qty', axis=1, level=1)
out_percentage2 = (1 - (sold_qty2 / offer_qty2)) * 100

# Handle division by zero (if Offer_Qty is 0)
out_percentage2 = out_percentage2.fillna(0)

# Add 'Out%' back to pivot_df
for garden in out_percentage2.columns:
    pivot_df2[(garden, 'Out%')] = out_percentage2[garden]

# Sort columns for better readability
pivot_df2 = pivot_df2.sort_index(axis=1)

# Checking the performance of 'Out%'
pivot_df2.loc[:, pivot_df2.columns.get_level_values(1) == 'Out%'].mean()

################################################
sum_cols = ['Sold_Qty', 'Grade%']
avg_cols = ['Avg_Price']
weight_col = 'Sold_Qty'

def add_subtotals(df):
    subtotals = []
    
    for category in ['PRIMARY', 'SECONDARY']:
        category_rows = df.loc[category]
        subtotal = pd.DataFrame(index=[(category, 'Subtotal')], columns=df.columns)
        
        for garden in df.columns.levels[0]:  
            # Sum Sold_Qty & Grade%
            subtotal[(garden, 'Sold_Qty')] = category_rows[(garden, 'Sold_Qty')].sum()
            subtotal[(garden, 'Grade%')] = category_rows[(garden, 'Grade%')].sum()
            
            # Weighted Average for Avg_Price
            total_weight = category_rows[(garden, weight_col)].sum()
            if total_weight > 0:
                subtotal[(garden, 'Avg_Price')] = (
                    (category_rows[(garden, 'Avg_Price')] * category_rows[(garden, weight_col)]).sum() / total_weight
                )
            else:
                subtotal[(garden, 'Avg_Price')] = 0

            # Sum Offer_Qty (Needed for Out% Calculation)
            subtotal[(garden, 'Offer_Qty')] = category_rows[(garden, 'Offer_Qty')].sum()

            # Correct Out% Calculation
            if subtotal[(garden, 'Offer_Qty')].iloc[0] > 0:
                subtotal[(garden, 'Out%')]=(1 -(subtotal[(garden,'Sold_Qty')] / subtotal[(garden,'Offer_Qty')]))*100
            else:
                subtotal[(garden, 'Out%')] = 0  # Avoid division by zero
        
        subtotals.append(subtotal)

    df = pd.concat([df] + subtotals).sort_index()

    # Compute Grand Total
    grand_total = pd.DataFrame(index=[('Grand Total', '')], columns=df.columns)

    for garden in df.columns.levels[0]:  
        grand_total[(garden, 'Sold_Qty')] = df.xs('Subtotal', level=1).loc[:, (garden, 'Sold_Qty')].sum()
        grand_total[(garden, 'Grade%')] = df.xs('Subtotal', level=1).loc[:, (garden, 'Grade%')].sum()
        grand_total[(garden, 'Offer_Qty')] = df.xs('Subtotal', level=1).loc[:, (garden, 'Offer_Qty')].sum()

        total_weight = df.xs('Subtotal', level=1).loc[:, (garden, weight_col)].sum()
        if total_weight > 0:
            grand_total[(garden, 'Avg_Price')] = (
                (df.xs('Subtotal', level=1).loc[:, (garden, 'Avg_Price')] * df.xs('Subtotal', level=1).loc[:, (garden, weight_col)]).sum()
                / total_weight
            )
        else:
            grand_total[(garden, 'Avg_Price')] = 0
        
        # Correct Out% Calculation for Grand Total
        if grand_total[(garden, 'Offer_Qty')].iloc[0] > 0:
            grand_total[(garden, 'Out%')] = (1 - (grand_total[(garden, 'Sold_Qty')] / grand_total[(garden, 'Offer_Qty')])) * 100
        else:
            grand_total[(garden, 'Out%')] = 0

    df = pd.concat([df, grand_total])
    
    return df

# Apply function
pivot_df2 = add_subtotals(pivot_df2)

# Custom sorting function

desired_order = ['Sold_Qty', 'Grade%', 'Avg_Price', 'Out%']

# Rearrange columns under each Garden name
new_columns = []
for garden in pivot_df2.columns.levels[0]:  # Iterate over the garden names
    for metric in desired_order:  # Maintain the desired metric order
        new_columns.append((garden, metric))

# Update the DataFrame with the new column order
pivot_df2 = pivot_df2[new_columns]

#########################Grade Sequence####################
desired_order2 = [
    "BOPL", "BPS", "BOP", "BOPSM", "BP", "OF", "PF", "PD", "D", "CD",
    "BOPL1", "BPS1", "BOP1", "BOPSM1", "BP1", "OF1","PF1", "PD1", "D1", "CD1"]

def reorder_group(group):
    # Split "Subtotal" rows from others
    subtotal = group[group.index.get_level_values(1) == 'Subtotal']
    others = group[group.index.get_level_values(1) != 'Subtotal']
    
    others = others.reindex(
        sorted(others.index,
            key=lambda x: desired_order2.index(x[1]) if x[1] in desired_order2 else len(desired_order2)
        ))
    
    # Append "Subtotal" back at the end
    return pd.concat([others, subtotal])

# Separate "Grand Total" row
grand_total2 = pivot_df2[pivot_df2.index.get_level_values(0) == 'Grand Total']
pivot_df_without_grand_total2 = pivot_df2[pivot_df2.index.get_level_values(0) != 'Grand Total']

pivot_df_reordered2 = pivot_df_without_grand_total2.groupby(level=0, group_keys=False).apply(reorder_group)

# Append "Grand Total" back at the bottom
pivot_df_final2 = pd.concat([pivot_df_reordered2, grand_total2])


#------------------------------------------DESIGN--------------------------------------------#

with pd.ExcelWriter("final_result_new2.xlsx", engine="openpyxl") as writer:
    pivot_df_final2.to_excel(writer, sheet_name="EST", startrow=0)
    
    start_row_2 = len(pivot_df_final2) + 6  # Ensure the second table starts at the correct position
    pivot_df_final.to_excel(writer, sheet_name="EST", startrow=start_row_2)

    workbook = writer.book
    worksheet = writer.sheets["EST"]
        
    worksheet.freeze_panes = "A3"

    # Define border, font, and colors
    thin_border = Border(
        left=Side(style="thin"),right=Side(style="thin"),
        top=Side(style="thin"),bottom=Side(style="thin")
    )
    bold_font = Font(bold=True)
    light_green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    light_blue_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    for cell in worksheet[1]:  # Row 1
        cell.fill = yellow_fill

    # Apply formatting for both tables
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
        first_cell = row[0]

        # Highlight "Subtotal" and "Grand Total" rows
        if first_cell.value and any(x in str(first_cell.value).strip() for x in ["Subtotal", "Grand Total"]):
            for cell in row:
                cell.font = bold_font
                cell.fill = light_green_fill

        # Apply border, alignment, and zero-value handling
        if first_cell.row < start_row_2 - 2 or first_cell.row >= start_row_2:  # Avoid formatting empty rows
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    if cell.value == 0:
                        cell.value = ""
                    else:
                        cell.number_format = "#,##,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
    for row in [20, 27]:
        for cell in worksheet[row]:
            cell.border = None
    
    m1=df1['SaleAlies'].max()

    # Rename headers for first two columns
    worksheet["A2"] = "SubTeaType"
    worksheet["A19"] = "SubTeaType"
    worksheet["B2"] = "Grade"
    worksheet["B19"] = "Grade"
    worksheet["A2"].font = bold_font
    worksheet["B2"].font = bold_font
    worksheet["A19"].font = bold_font
    worksheet["B19"].font = bold_font
    
    worksheet["A1"] = f"For Sale {m1-52 if m1>52 else m1}"
    worksheet["A1"].font = Font(bold=True, color="FF0000")
    
    worksheet["A20"] = f"Upto Sale {m1-52 if m1>52 else m1}"
    worksheet["A20"].font = Font(bold=True, color="FF0000")
    
    # Adjust column width
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 15

    # Apply light blue fill for headers (Fix for both tables)
    for row in [2, start_row_2 + 1]:  # Apply header formatting for both tables
        for cell in worksheet.iter_rows(min_row=row, max_row=row, min_col=3, max_col=worksheet.max_column):
            for header_cell in cell:
                header_cell.fill = light_blue_fill  

    # Ensure subtotals are bold
    def apply_bold_to_subtotals(worksheet):
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == 'Subtotal':
                    for sub_cell in row:
                        sub_cell.font = bold_font

    apply_bold_to_subtotals(worksheet)
